from __future__ import annotations

import argparse
import math
import re
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
KPI_FILL = PatternFill("solid", fgColor="EAF4E2")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=14)
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Gera um workbook Excel analitico a partir de arquivos agregados "
            "do AvaliaRJ/AlfabetizaRJ por escola e por turma."
        )
    )
    parser.add_argument(
        "--arquivo-escola",
        required=True,
        help="Caminho do arquivo agregado por escola (.xls, .xlsx ou .csv).",
    )
    parser.add_argument(
        "--arquivo-turma",
        required=True,
        help="Caminho do arquivo agregado por turma e escola (.xls, .xlsx ou .csv).",
    )
    parser.add_argument(
        "--saida",
        required=True,
        help="Caminho do workbook .xlsx de saida.",
    )
    parser.add_argument(
        "--arquivo-legenda",
        default="export/generated/legenda_habilidades_alfabetizarj_2o_captura_2025.csv",
        help=(
            "CSV opcional com legenda/crosswalk de habilidades Hxx. "
            "Quando ausente, o workbook usa apenas H01..Hxx."
        ),
    )
    return parser.parse_args()


def repair_text(value: object) -> object:
    if not isinstance(value, str):
        return value

    text = value.replace("\xa0", " ").strip()
    if any(token in text for token in ("Ã", "Â", "�")):
        try:
            text = text.encode("latin1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass

    text = text.replace("Â ", " ").replace("Â", "")
    return " ".join(text.split())


def read_input(path_str: str) -> pd.DataFrame:
    path = Path(path_str)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        df = pd.read_csv(path)
    elif suffix == ".xls":
        df = pd.read_excel(path, engine="xlrd")
    elif suffix == ".xlsx":
        df = pd.read_excel(path, engine="openpyxl")
    else:
        raise ValueError(f"Formato nao suportado: {path}")

    df.columns = [repair_text(col) for col in df.columns]
    for column in df.columns:
        if df[column].dtype == object:
            df[column] = df[column].map(repair_text)
    return df


def as_ratio(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.dropna().empty:
        return numeric
    if numeric.max() > 1.5:
        return numeric / 100.0
    return numeric


def as_percent(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.dropna().empty:
        return numeric
    if numeric.max() <= 1.5:
        return numeric * 100.0
    return numeric


def normalize_skill_code(column_name: str) -> str:
    match = re.search(r"H\s*(\d+)", column_name, flags=re.IGNORECASE)
    if not match:
        return column_name
    return f"H{int(match.group(1)):02d}"


def weighted_average(values: pd.Series, weights: pd.Series) -> float:
    mask = values.notna() & weights.notna()
    if not mask.any():
        return math.nan
    return float(np.average(values[mask], weights=weights[mask]))


def load_legend(path_str: str, component: str) -> pd.DataFrame:
    path = Path(path_str)
    if not path.exists():
        return pd.DataFrame(columns=["habilidade", "codigo_referencia_publica", "descricao_origem", "status_legenda"])

    legend = pd.read_csv(path)
    legend.columns = [repair_text(col) for col in legend.columns]
    for column in legend.columns:
        if legend[column].dtype == object:
            legend[column] = legend[column].map(repair_text)

    legend["habilidade"] = legend["codigo_origem"].map(normalize_skill_code)
    return legend.loc[legend["componente"] == component].copy()


def infer_component(df: pd.DataFrame) -> str:
    if "Componente Curricular" not in df.columns or df.empty:
        return "LP"
    component = str(df["Componente Curricular"].iloc[0]).strip().upper()
    return component if component else "LP"


def prepare_base(df: pd.DataFrame, include_turma: bool) -> tuple[pd.DataFrame, list[str], list[str]]:
    df = df.copy()

    ratio_columns = [
        "Participação(%)",
        "Abaixo do básico",
        "Básico",
        "Adequado",
        "Avançado",
    ]
    numeric_columns = ["Previstos", "Avaliados", "Proficiência Média"]
    skill_columns = [column for column in df.columns if re.fullmatch(r"H\s*\d+\s*\(%\)", column, flags=re.IGNORECASE)]

    for column in numeric_columns + skill_columns:
        if column in df.columns:
            df[column] = pd.to_numeric(df[column], errors="coerce")
    for column in ratio_columns:
        if column in df.columns:
            df[column] = as_ratio(df[column])

    df["Participacao %"] = df["Participação(%)"] * 100.0
    df["Abaixo do básico %"] = df["Abaixo do básico"] * 100.0
    df["Básico %"] = df["Básico"] * 100.0
    df["Adequado %"] = df["Adequado"] * 100.0
    df["Avançado %"] = df["Avançado"] * 100.0
    df["Crítico (%)"] = df["Abaixo do básico %"] + df["Básico %"]
    df["Esperado+ (%)"] = df["Adequado %"] + df["Avançado %"]

    for column in skill_columns:
        df[column] = as_percent(df[column])

    id_columns = [
        "Avaliação",
        "Rede",
        "Ano Escolar",
        "Componente Curricular",
        "Estado",
        "Regional",
        "Município",
        "Escola",
    ]
    if include_turma:
        id_columns.extend(["Código da Turma", "Turma"])

    present_id_columns = [column for column in id_columns if column in df.columns]
    metric_columns = ["Previstos", "Avaliados", "Participacao %", "Proficiência Média"]
    level_columns = ["Abaixo do básico %", "Básico %", "Adequado %", "Avançado %", "Crítico (%)", "Esperado+ (%)"]

    report = df[present_id_columns + metric_columns + level_columns + skill_columns].copy()
    return report, skill_columns, level_columns


def best_and_worst_skills(row: pd.Series, skill_columns: list[str], labels: dict[str, str]) -> tuple[str, str]:
    skill_values = {labels[column]: row[column] for column in skill_columns if pd.notna(row[column])}
    if not skill_values:
        return "", ""
    weakest = min(skill_values, key=skill_values.get)
    strongest = max(skill_values, key=skill_values.get)
    return weakest, strongest


def build_ranked_reports(
    school_df: pd.DataFrame,
    turma_df: pd.DataFrame,
    skill_columns: list[str],
    skill_labels: dict[str, str],
) -> tuple[pd.DataFrame, pd.DataFrame, float]:
    network_prof = weighted_average(school_df["Proficiência Média"], school_df["Avaliados"])

    school_report = school_df.copy().sort_values(["Proficiência Média", "Avaliados"], ascending=[False, False]).reset_index(drop=True)
    school_report.insert(0, "Ranking", school_report.index + 1)
    school_report["Proficiência vs rede"] = school_report["Proficiência Média"] - network_prof
    school_report[["Habilidade mais frágil", "Habilidade mais forte"]] = school_report.apply(
        lambda row: pd.Series(best_and_worst_skills(row, skill_columns, skill_labels)),
        axis=1,
    )

    turma_report = turma_df.copy().sort_values(["Proficiência Média", "Avaliados"], ascending=[False, False]).reset_index(drop=True)
    turma_report.insert(0, "Ranking", turma_report.index + 1)
    turma_report["Proficiência vs rede"] = turma_report["Proficiência Média"] - network_prof
    turma_report[["Habilidade mais frágil", "Habilidade mais forte"]] = turma_report.apply(
        lambda row: pd.Series(best_and_worst_skills(row, skill_columns, skill_labels)),
        axis=1,
    )

    return school_report, turma_report, network_prof


def build_skill_summary(
    school_df: pd.DataFrame,
    turma_df: pd.DataFrame,
    skill_columns: list[str],
    skill_labels: dict[str, str],
    legend: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    legend_map = legend.set_index("habilidade").to_dict("index")
    summary_rows = []
    for column in skill_columns:
        skill_code = normalize_skill_code(column)
        legend_item = legend_map.get(skill_code, {})
        average_value = weighted_average(school_df[column], school_df["Avaliados"])
        if average_value >= 80:
            priority = "Consolidada"
        elif average_value >= 60:
            priority = "Acompanhar"
        else:
            priority = "Prioridade alta"
        summary_rows.append(
            {
                "Habilidade": skill_labels[column],
                "Codigo publico": legend_item.get("codigo_referencia_publica", ""),
                "Descricao": legend_item.get("descricao_origem", ""),
                "Status legenda": legend_item.get("status_legenda", ""),
                "Media da rede (%)": round(average_value, 2),
                "Prioridade": priority,
            }
        )

    skill_summary = pd.DataFrame(summary_rows).sort_values("Media da rede (%)").reset_index(drop=True)

    school_skill_columns = ["Escola", "Avaliados"] + skill_columns
    school_skills = school_df[school_skill_columns].copy()
    school_skills = school_skills.rename(columns={column: skill_labels[column] for column in skill_columns})
    school_skills["Media habilidades (%)"] = school_skills[[skill_labels[column] for column in skill_columns]].mean(axis=1)
    school_skills = school_skills.sort_values("Media habilidades (%)", ascending=False).reset_index(drop=True)

    turma_skill_columns = ["Escola", "Turma", "Avaliados"] + skill_columns
    turma_skills = turma_df[turma_skill_columns].copy()
    turma_skills = turma_skills.rename(columns={column: skill_labels[column] for column in skill_columns})
    turma_skills["Media habilidades (%)"] = turma_skills[[skill_labels[column] for column in skill_columns]].mean(axis=1)
    turma_skills = turma_skills.sort_values("Media habilidades (%)", ascending=False).reset_index(drop=True)

    return skill_summary, school_skills, turma_skills


def build_consistency_check(
    school_raw: pd.DataFrame,
    turma_raw: pd.DataFrame,
    skill_columns: list[str],
) -> pd.DataFrame:
    grouped = turma_raw.groupby("Escola", dropna=False)
    rows = []
    for _, school_row in school_raw.iterrows():
        school_name = school_row["Escola"]
        chunk = grouped.get_group(school_name)
        prof_turma = weighted_average(chunk["Proficiência Média"], chunk["Avaliados"])
        part_turma = chunk["Avaliados"].sum() / chunk["Previstos"].sum() * 100.0
        row = {
            "Escola": school_name,
            "Avaliados arquivo escola": int(school_row["Avaliados"]),
            "Avaliados soma turmas": int(chunk["Avaliados"].sum()),
            "Participacao escola (%)": round(float(school_row["Participacao %"]), 2),
            "Participacao turmas (%)": round(float(part_turma), 2),
            "Delta participacao (p.p.)": round(float(school_row["Participacao %"] - part_turma), 2),
            "Proficiencia escola": round(float(school_row["Proficiência Média"]), 2),
            "Proficiencia turmas": round(float(prof_turma), 2),
            "Delta proficiencia": round(float(school_row["Proficiência Média"] - prof_turma), 2),
        }
        for column in skill_columns:
            row[f"{normalize_skill_code(column)} delta"] = round(
                float(school_row[column] - weighted_average(chunk[column], chunk["Avaliados"])),
                2,
            )
        rows.append(row)

    return pd.DataFrame(rows)


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 40)


def style_table_sheet(worksheet, table_name: str, percentage_columns: set[str], decimal_columns: set[str]) -> None:
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    if max_row < 2 or max_col < 1:
        return

    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.freeze_panes = "A2"

    headers = [cell.value for cell in worksheet[1]]
    for cell in worksheet[1]:
        cell.fill = TITLE_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    header_index = {header: index + 1 for index, header in enumerate(headers)}
    for header, column_index in header_index.items():
        for row_index in range(2, max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = THIN_BORDER
            if header in percentage_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "0.0"
            elif header in decimal_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    autosize_columns(worksheet)


def add_heatmap(worksheet, first_column: int, last_column: int) -> None:
    if worksheet.max_row < 2 or last_column < first_column:
        return
    worksheet.conditional_formatting.add(
        f"{get_column_letter(first_column)}2:{get_column_letter(last_column)}{worksheet.max_row}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )


def build_summary_sheet(
    writer: pd.ExcelWriter,
    metadata: dict[str, str],
    school_report: pd.DataFrame,
    turma_report: pd.DataFrame,
    skill_summary: pd.DataFrame,
    network_prof: float,
) -> None:
    workbook = writer.book
    if "Resumo Executivo" in workbook.sheetnames:
        del workbook["Resumo Executivo"]
    worksheet = workbook.create_sheet("Resumo Executivo", 0)

    worksheet["A1"] = "Analise Pontual - AvaliaRJ 2025"
    worksheet["A1"].fill = TITLE_FILL
    worksheet["A1"].font = WHITE_FONT
    worksheet["A1"].alignment = Alignment(horizontal="left")
    worksheet.merge_cells("A1:H1")

    worksheet["A3"] = "Contexto"
    worksheet["A3"].fill = SECTION_FILL
    worksheet["A3"].font = SECTION_FONT
    worksheet.merge_cells("A3:B3")

    context_rows = [
        ("Avaliacao", metadata.get("Avaliação", "")),
        ("Rede", metadata.get("Rede", "")),
        ("Ano escolar", metadata.get("Ano Escolar", "")),
        ("Componente", metadata.get("Componente Curricular", "")),
        ("Municipio", metadata.get("Município", "")),
        ("Regional", metadata.get("Regional", "")),
    ]
    current_row = 4
    for label, value in context_rows:
        worksheet[f"A{current_row}"] = label
        worksheet[f"B{current_row}"] = value
        current_row += 1

    total_previstos = int(school_report["Previstos"].sum())
    total_avaliados = int(school_report["Avaliados"].sum())
    participation = total_avaliados / total_previstos if total_previstos else math.nan
    critical_share = weighted_average(school_report["Crítico (%)"], school_report["Avaliados"])
    expected_share = weighted_average(school_report["Esperado+ (%)"], school_report["Avaliados"])

    worksheet["D3"] = "KPIs da rede"
    worksheet["D3"].fill = SECTION_FILL
    worksheet["D3"].font = SECTION_FONT
    worksheet.merge_cells("D3:F3")

    kpis = [
        ("Escolas", int(school_report["Escola"].nunique()), "0"),
        ("Turmas", int(turma_report["Turma"].nunique()), "0"),
        ("Previstos", total_previstos, "0"),
        ("Avaliados", total_avaliados, "0"),
        ("Participacao", participation, "0.0%"),
        ("Proficiencia media", network_prof, "0.00"),
        ("Critico", critical_share / 100.0, "0.0%"),
        ("Esperado+", expected_share / 100.0, "0.0%"),
    ]
    for index, (label, value, number_format) in enumerate(kpis):
        column = "D" if index < 4 else "E"
        row_base = 4 + (index % 4) * 2
        worksheet[f"{column}{row_base}"] = label
        worksheet[f"{column}{row_base}"].fill = KPI_FILL
        worksheet[f"{column}{row_base}"].font = Font(bold=True)
        worksheet[f"{column}{row_base + 1}"] = value
        worksheet[f"{column}{row_base + 1}"].number_format = number_format

    best_school = school_report.iloc[0]
    weakest_school = school_report.iloc[-1]
    best_turma = turma_report.iloc[0]
    weakest_turma = turma_report.iloc[-1]
    priority_skills = skill_summary.head(3)
    consolidated_skills = skill_summary.tail(3).sort_values("Media da rede (%)", ascending=False)

    worksheet["A11"] = "Principais achados"
    worksheet["A11"].fill = SECTION_FILL
    worksheet["A11"].font = SECTION_FONT
    worksheet.merge_cells("A11:H11")

    findings = [
        (
            "1",
            (
                f"Rede com {total_avaliados} avaliados de {total_previstos} previstos "
                f"({participation:.1%}) e proficiencia media ponderada de {network_prof:.2f}."
            ),
        ),
        (
            "2",
            (
                f"Melhor escola em proficiencia: {best_school['Escola']} "
                f"({best_school['Proficiência Média']:.0f}). "
                f"Escola com menor proficiencia: {weakest_school['Escola']} "
                f"({weakest_school['Proficiência Média']:.0f})."
            ),
        ),
        (
            "3",
            (
                f"Turma destaque: {best_turma['Turma']} / {best_turma['Escola']} "
                f"({best_turma['Proficiência Média']:.0f}). "
                f"Turma mais critica: {weakest_turma['Turma']} / {weakest_turma['Escola']} "
                f"({weakest_turma['Proficiência Média']:.0f})."
            ),
        ),
        (
            "4",
            (
                "Habilidades com menor dominio na rede: "
                + ", ".join(
                    f"{row['Habilidade']} ({row['Media da rede (%)']:.1f}%)"
                    for _, row in priority_skills.iterrows()
                )
                + "."
            ),
        ),
        (
            "5",
            (
                "Habilidades mais consolidadas: "
                + ", ".join(
                    f"{row['Habilidade']} ({row['Media da rede (%)']:.1f}%)"
                    for _, row in consolidated_skills.iterrows()
                )
                + "."
            ),
        ),
    ]

    row_pointer = 12
    for marker, text in findings:
        worksheet[f"A{row_pointer}"] = marker
        worksheet[f"B{row_pointer}"] = text
        worksheet[f"B{row_pointer}"].alignment = Alignment(wrap_text=True, vertical="top")
        row_pointer += 1

    worksheet["A19"] = "Observacoes"
    worksheet["A19"].fill = SECTION_FILL
    worksheet["A19"].font = SECTION_FONT
    worksheet.merge_cells("A19:H19")

    notes = [
        "Analise offline e pontual, sem persistencia de microdados na base canonica do projeto.",
        "As habilidades Hxx usam legenda documental de captura 2025 quando disponivel.",
        "Participacao e distribuicao por nivel foram ponderadas por avaliados.",
    ]
    for offset, note in enumerate(notes, start=20):
        worksheet[f"A{offset}"] = "- " + note
        worksheet[f"A{offset}"].alignment = Alignment(wrap_text=True)
        worksheet.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=8)

    worksheet["H3"] = f"Gerado em {datetime.now():%d/%m/%Y %H:%M}"
    worksheet["H3"].alignment = Alignment(horizontal="right")

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER

    worksheet.row_dimensions[1].height = 24
    autosize_columns(worksheet)


def add_charts(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    resumo = workbook["Resumo Executivo"]
    escolas = workbook["Escolas"]
    habilidades = workbook["Habilidades Rede"]

    prof_col = None
    escola_col = None
    for index, cell in enumerate(escolas[1], start=1):
        if cell.value == "Escola":
            escola_col = index
        if cell.value == "Proficiência Média":
            prof_col = index

    if prof_col and escola_col and escolas.max_row >= 2:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Proficiência por escola"
        chart.y_axis.title = "Escola"
        chart.x_axis.title = "Proficiência"
        data = Reference(escolas, min_col=prof_col, min_row=1, max_row=escolas.max_row)
        categories = Reference(escolas, min_col=escola_col, min_row=2, max_row=escolas.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 11
        resumo.add_chart(chart, "D12")

    habilidade_col = None
    media_col = None
    for index, cell in enumerate(habilidades[1], start=1):
        if cell.value == "Habilidade":
            habilidade_col = index
        if cell.value == "Media da rede (%)":
            media_col = index

    if habilidade_col and media_col and habilidades.max_row >= 2:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 12
        chart.title = "Habilidades da rede"
        chart.y_axis.title = "Habilidade"
        chart.x_axis.title = "Dominio (%)"
        data = Reference(habilidades, min_col=media_col, min_row=1, max_row=habilidades.max_row)
        categories = Reference(habilidades, min_col=habilidade_col, min_row=2, max_row=habilidades.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 11
        resumo.add_chart(chart, "D27")


def write_workbook(
    output_path: Path,
    metadata: dict[str, str],
    school_report: pd.DataFrame,
    turma_report: pd.DataFrame,
    skill_summary: pd.DataFrame,
    school_skills: pd.DataFrame,
    turma_skills: pd.DataFrame,
    consistency_df: pd.DataFrame,
    methodology_lines: list[str],
    network_prof: float,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        school_report.to_excel(writer, sheet_name="Escolas", index=False)
        turma_report.to_excel(writer, sheet_name="Turmas", index=False)
        skill_summary.to_excel(writer, sheet_name="Habilidades Rede", index=False)
        school_skills.to_excel(writer, sheet_name="Habilidades Escola", index=False)
        turma_skills.to_excel(writer, sheet_name="Habilidades Turma", index=False)
        consistency_df.to_excel(writer, sheet_name="Conferencia", index=False)

        workbook = writer.book
        methodology_sheet = workbook.create_sheet("Metodologia")
        methodology_sheet["A1"] = "Metodologia"
        methodology_sheet["A1"].fill = TITLE_FILL
        methodology_sheet["A1"].font = WHITE_FONT
        methodology_sheet.merge_cells("A1:H1")
        for row_index, line in enumerate(methodology_lines, start=3):
            methodology_sheet[f"A{row_index}"] = line
            methodology_sheet[f"A{row_index}"].alignment = Alignment(wrap_text=True)
            methodology_sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=8)
        autosize_columns(methodology_sheet)

        build_summary_sheet(writer, metadata, school_report, turma_report, skill_summary, network_prof)

        percentage_headers = {
            "Participacao %",
            "Abaixo do básico %",
            "Básico %",
            "Adequado %",
            "Avançado %",
            "Crítico (%)",
            "Esperado+ (%)",
            "Media da rede (%)",
            "Media habilidades (%)",
        }
        percentage_headers.update({column for column in school_skills.columns if "(D" in str(column) or str(column).startswith("H")})
        percentage_headers.update({column for column in turma_skills.columns if "(D" in str(column) or str(column).startswith("H")})

        style_table_sheet(workbook["Escolas"], "TblEscolasAnalise", percentage_headers, {"Proficiência vs rede"})
        style_table_sheet(workbook["Turmas"], "TblTurmasAnalise", percentage_headers, {"Proficiência vs rede"})
        style_table_sheet(workbook["Habilidades Rede"], "TblHabilidadesRede", {"Media da rede (%)"}, set())
        style_table_sheet(workbook["Habilidades Escola"], "TblHabilidadesEscola", percentage_headers, set())
        style_table_sheet(workbook["Habilidades Turma"], "TblHabilidadesTurma", percentage_headers, set())
        style_table_sheet(workbook["Conferencia"], "TblConferencia", set(), {"Delta participacao (p.p.)", "Delta proficiencia"})

        escolas_headers = [cell.value for cell in workbook["Escolas"][1]]
        turmas_headers = [cell.value for cell in workbook["Turmas"][1]]
        habilidades_escola_headers = [cell.value for cell in workbook["Habilidades Escola"][1]]
        habilidades_turma_headers = [cell.value for cell in workbook["Habilidades Turma"][1]]

        if "Proficiência Média" in escolas_headers:
            add_heatmap(
                workbook["Escolas"],
                escolas_headers.index("Proficiência Média") + 1,
                escolas_headers.index("Esperado+ (%)") + 1,
            )
        if "Proficiência Média" in turmas_headers:
            add_heatmap(
                workbook["Turmas"],
                turmas_headers.index("Proficiência Média") + 1,
                turmas_headers.index("Esperado+ (%)") + 1,
            )

        first_skill_column_school = 3
        if "Avaliados" in habilidades_escola_headers:
            first_skill_column_school = habilidades_escola_headers.index("Avaliados") + 2
        add_heatmap(workbook["Habilidades Escola"], first_skill_column_school, workbook["Habilidades Escola"].max_column)

        first_skill_column_turma = 4
        if "Avaliados" in habilidades_turma_headers:
            first_skill_column_turma = habilidades_turma_headers.index("Avaliados") + 2
        add_heatmap(workbook["Habilidades Turma"], first_skill_column_turma, workbook["Habilidades Turma"].max_column)

        add_charts(writer)


def main() -> None:
    args = parse_args()

    school_raw = read_input(args.arquivo_escola)
    turma_raw = read_input(args.arquivo_turma)

    component = infer_component(school_raw)
    legend = load_legend(args.arquivo_legenda, component)

    school_report_base, skill_columns_school, _ = prepare_base(school_raw, include_turma=False)
    turma_report_base, skill_columns_turma, _ = prepare_base(turma_raw, include_turma=True)
    skill_columns = [column for column in skill_columns_school if column in skill_columns_turma]
    if not skill_columns:
        raise ValueError("Nenhuma coluna de habilidade Hxx foi encontrada em comum entre os arquivos.")

    legend_map = legend.set_index("habilidade").to_dict("index")
    skill_labels = {}
    for column in skill_columns:
        skill_code = normalize_skill_code(column)
        code = legend_map.get(skill_code, {}).get("codigo_referencia_publica", "")
        skill_labels[column] = f"{skill_code} ({code})" if code else skill_code

    school_report, turma_report, network_prof = build_ranked_reports(
        school_report_base,
        turma_report_base,
        skill_columns,
        skill_labels,
    )
    skill_summary, school_skills, turma_skills = build_skill_summary(
        school_report_base,
        turma_report_base,
        skill_columns,
        skill_labels,
        legend,
    )
    consistency_df = build_consistency_check(school_report_base, turma_report_base, skill_columns)

    metadata_columns = [
        "Avaliação",
        "Rede",
        "Ano Escolar",
        "Componente Curricular",
        "Município",
        "Regional",
    ]
    metadata = {column: str(school_raw.iloc[0][column]) for column in metadata_columns if column in school_raw.columns and not school_raw.empty}

    methodology_lines = [
        f"Arquivo por escola: {Path(args.arquivo_escola).resolve()}",
        f"Arquivo por turma: {Path(args.arquivo_turma).resolve()}",
        f"Legenda de habilidades: {Path(args.arquivo_legenda).resolve() if Path(args.arquivo_legenda).exists() else 'nao utilizada'}",
        "A saida consolida os dados em um workbook Excel para leitura gerencial.",
        "Nao ha gravacao de microdados em banco; a analise e feita localmente com agregacao em memoria.",
        "Participacao global = avaliados / previstos.",
        "Proficiencia media da rede = media ponderada por avaliados.",
        "Niveis de desempenho e habilidades tambem usam ponderacao por avaliados.",
        "A aba Conferencia compara o arquivo por escola com a reconstituicao a partir das turmas.",
        "Quando houver legenda Hxx disponivel, o workbook exibe o codigo publico associado (por exemplo D02).",
    ]

    write_workbook(
        Path(args.saida),
        metadata,
        school_report,
        turma_report,
        skill_summary,
        school_skills,
        turma_skills,
        consistency_df,
        methodology_lines,
        network_prof,
    )

    print(f"Workbook gerado com sucesso em: {Path(args.saida).resolve()}")


if __name__ == "__main__":
    main()
